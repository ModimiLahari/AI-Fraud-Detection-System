import io
import json
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
)
from reportlab.lib.units import cm
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app import models


def build_fraud_pdf_report(customer: models.Customer, report: models.FraudReport) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2 * cm, bottomMargin=2 * cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleX", parent=styles["Title"], textColor=colors.HexColor("#0F172A"))
    heading_style = ParagraphStyle("HeadingX", parent=styles["Heading2"], textColor=colors.HexColor("#1D4ED8"))

    elements = []
    elements.append(Paragraph("Bank Fraud & Risk Report", title_style))
    elements.append(Spacer(1, 12))

    info_data = [
        ["Customer Name", customer.full_name],
        ["Customer Code", customer.customer_code],
        ["Branch", customer.branch],
        ["Risk Score", f"{report.risk_score} / 100"],
        ["Risk Level", report.risk_level],
        ["Generated At", report.generated_at.strftime("%d-%b-%Y %H:%M")],
    ]
    info_table = Table(info_data, colWidths=[5 * cm, 10 * cm])
    info_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#E2E8F0")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("PADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 20))

    elements.append(Paragraph("Triggered Risk Rules", heading_style))
    elements.append(Spacer(1, 8))

    rules = json.loads(report.triggered_rules) if report.triggered_rules else []
    if rules:
        rule_rows = [["Rule", "Points", "Reason"]]
        for r in rules:
            rule_rows.append([r["rule"], str(r["points"]), Paragraph(r["reason"], styles["BodyText"])])
        rule_table = Table(rule_rows, colWidths=[4 * cm, 2 * cm, 9 * cm])
        rule_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1D4ED8")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("PADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(rule_table)
    else:
        elements.append(Paragraph("No risk rules were triggered.", styles["BodyText"]))

    elements.append(Spacer(1, 20))
    elements.append(Paragraph("AI Explanation", heading_style))
    elements.append(Spacer(1, 8))
    elements.append(Paragraph(report.ai_explanation or "Not available", styles["BodyText"]))

    elements.append(Spacer(1, 20))
    elements.append(Paragraph("Recommended Actions", heading_style))
    elements.append(Spacer(1, 8))
    actions = json.loads(report.recommended_actions) if report.recommended_actions else []
    for a in actions:
        elements.append(Paragraph(f"• {a}", styles["BodyText"]))

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def build_fraud_excel_report(customers_with_reports) -> bytes:
    """
    customers_with_reports: list of tuples (Customer, FraudReport or None)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Fraud Risk Report"

    headers = ["Customer Code", "Name", "Branch", "Risk Score", "Risk Level", "Generated At", "Top Reason"]
    ws.append(headers)
    header_fill = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill

    for customer, report in customers_with_reports:
        top_reason = ""
        if report and report.triggered_rules:
            rules = json.loads(report.triggered_rules)
            if rules:
                top_reason = rules[0]["reason"]
        ws.append([
            customer.customer_code,
            customer.full_name,
            customer.branch,
            report.risk_score if report else 0,
            report.risk_level if report else "Not Evaluated",
            report.generated_at.strftime("%d-%b-%Y %H:%M") if report else "",
            top_reason,
        ])

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
