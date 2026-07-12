"""
Portfolio-wide Excel export — one workbook, two sheets:
  1. "Customers"  — every customer with risk score, tier (color-coded), branch, top rule
  2. "Alerts"     — recent alerts feed, for offline review by risk officers

Uses openpyxl only (no LibreOffice/pandas-excel-writer binary dependency —
keeps the deploy footprint small on Render/Railway free tier).
Returns raw .xlsx bytes.
"""

import io
import logging
from datetime import datetime
from typing import List, Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

logger = logging.getLogger("reports.excel_generator")

HEADER_FILL = PatternFill(start_color="0B1220", end_color="0B1220", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN_BORDER = Border(*(Side(style="thin", color="E2E8F0"),) * 4)

TIER_FILLS = {
    "Low": PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
    "Medium": PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid"),
    "High": PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid"),
    "Critical": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
}


def _risk_tier(score: int) -> str:
    if score >= 80:
        return "Critical"
    if score >= 60:
        return "High"
    if score >= 30:
        return "Medium"
    return "Low"


def _style_header_row(ws, row_idx: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _autofit(ws, max_width: int = 45) -> None:
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max_width, max(10, length + 2))


def build_portfolio_excel(
    customers: List[Dict[str, Any]],
    alerts: List[Dict[str, Any]] | None = None,
) -> bytes:
    """
    customers: [{"id","name","branch","loan_amount","risk_score","top_rule"}, ...]
    alerts: [{"id","customer_name","severity","message","created_at"}, ...]
    """
    alerts = alerts or []
    wb = Workbook()

    # --- Sheet 1: Customers ---
    ws1 = wb.active
    ws1.title = "Customers"
    headers = ["Customer ID", "Name", "Branch", "Loan Amount", "Risk Score", "Risk Tier", "Top Risk Factor"]
    ws1.append(headers)
    _style_header_row(ws1, 1, len(headers))

    for c in customers:
        score = int(c.get("risk_score", 0))
        tier = _risk_tier(score)
        row = [
            c.get("id", ""), c.get("name", ""), c.get("branch", ""),
            c.get("loan_amount", ""), score, tier, c.get("top_rule", "—"),
        ]
        ws1.append(row)
        r = ws1.max_row
        for col in range(1, len(headers) + 1):
            ws1.cell(row=r, column=col).border = THIN_BORDER
        ws1.cell(row=r, column=6).fill = TIER_FILLS.get(tier, PatternFill())

    if customers:
        table_ref = f"A1:{get_column_letter(len(headers))}{ws1.max_row}"
        table = Table(displayName="CustomersTable", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        ws1.add_table(table)
    _autofit(ws1)
    ws1.freeze_panes = "A2"

    # --- Sheet 2: Alerts ---
    ws2 = wb.create_sheet("Alerts")
    alert_headers = ["Alert ID", "Customer", "Severity", "Message", "Created At"]
    ws2.append(alert_headers)
    _style_header_row(ws2, 1, len(alert_headers))

    severity_fills = {
        "critical": TIER_FILLS["Critical"], "high": TIER_FILLS["High"],
        "medium": TIER_FILLS["Medium"], "low": TIER_FILLS["Low"],
    }
    for a in alerts:
        row = [
            a.get("id", ""), a.get("customer_name", ""),
            str(a.get("severity", "")).title(), a.get("message", ""),
            a.get("created_at", ""),
        ]
        ws2.append(row)
        r = ws2.max_row
        for col in range(1, len(alert_headers) + 1):
            ws2.cell(row=r, column=col).border = THIN_BORDER
        sev = str(a.get("severity", "")).lower()
        if sev in severity_fills:
            ws2.cell(row=r, column=3).fill = severity_fills[sev]

    if alerts:
        table_ref = f"A1:{get_column_letter(len(alert_headers))}{ws2.max_row}"
        table2 = Table(displayName="AlertsTable", ref=table_ref)
        table2.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        ws2.add_table(table2)
    _autofit(ws2)
    ws2.freeze_panes = "A2"

    # --- Summary sheet ---
    ws3 = wb.create_sheet("Summary", 0)
    ws3.append(["Fraud Portfolio Summary"])
    ws3["A1"].font = Font(bold=True, size=14, color="0B1220")
    ws3.append([f"Generated: {datetime.now().strftime('%d %b %Y, %H:%M')}"])
    ws3.append([])
    tier_counts = {"Low": 0, "Medium": 0, "High": 0, "Critical": 0}
    for c in customers:
        tier_counts[_risk_tier(int(c.get("risk_score", 0)))] += 1
    ws3.append(["Risk Tier", "Customer Count"])
    _style_header_row(ws3, 4, 2)
    for tier, count in tier_counts.items():
        ws3.append([tier, count])
        ws3.cell(row=ws3.max_row, column=1).fill = TIER_FILLS[tier]
    ws3.append([])
    ws3.append(["Total Customers", len(customers)])
    ws3.append(["Total Alerts", len(alerts)])
    _autofit(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()
    buf.close()
    logger.info("Generated portfolio Excel export (%d customers, %d alerts, %d bytes)",
                len(customers), len(alerts), len(xlsx_bytes))
    return xlsx_bytes
